import os

from django.core.management import BaseCommand, CommandError
from django.conf import settings
from openpyxl import load_workbook

from basedata.models import ReportType, AccountingSubject


class Command(BaseCommand):
    help = '填充数据'

    def handle(self, *args, **options):
        wb = load_workbook(os.path.join(
            os.path.join(settings.BASE_DIR, 'fixtures'),
            'report_types.xlsx'))
        ws = wb['report_types']

        current_type = None
        parent_subjects = [None for i in range(ws.max_column)]
        for row in ws.iter_rows(min_row=1, max_col=ws.max_column, max_row=ws.max_row, values_only=True):
            for i in range(ws.max_column):
                if isinstance(row[i], str) and row[i].strip():
                    if i == 0:
                        current_type, created = ReportType.objects.get_or_create(name=row[i].strip())

                    else:
                        subject, created = AccountingSubject.objects.get_or_create(
                            name=row[i].strip(),
                            report_type=current_type,
                            parent=parent_subjects[i - 1]
                        )

                        parent_subjects[i] = subject

                    continue  # 一行只处理一个
